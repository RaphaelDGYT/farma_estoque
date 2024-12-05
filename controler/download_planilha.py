import os
import sys 
sys.path.append(os.path.join(os.path.dirname(__file__), ".."))
from model.relatorio import Relatorio
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
from pytz import timezone


class RelatorioExcel(Relatorio):
    def __init__(self):
        super().__init__()
        self.wb = Workbook()
        self.ws = self.wb.active
        self.fuso_horario = timezone("America/Sao_Paulo")
        self.cabecalho = {"A2": "Loja:", "A3": "Laboratório:", "A4": "Produto:", "E4": "Data:", "E6": "Horário:",
                        "B2": "Drogaria MatosFarma", "B3": "TODOS OS LABORATÓRIOS", "B4": "TODOS OS PRODUTOS",
                        "F4": datetime.now(self.fuso_horario).strftime("%d/%m/%Y"), "F6": datetime.now(self.fuso_horario).strftime("%H:%M:%S")}

        self.colunas = {"A10": "Código Produto", "B10": "Descrição Produto", "C10": "Laboratório",
                        "D10": "Lista/Adendo", "E10": "Lote", "F10": "Registro MS",
                        "G10": "Validade", "H10": "Estoque"}
        
    def planilha(self):
        # Adicionando dados da coluna "Registro MS" em uma lista para ser formatada
        tabela = super().relatorio()
        registro = []
        for i in tabela["Registro MS"]:
            formart_number = f"{i[:1]}.{i[1:5]}.{i[5:9]}.{i[9:12]}-{i[12:]}"
            registro.append(formart_number)

        for i in range(len(registro)):
            tabela.loc[i, "Registro MS"] = registro[i]

        # Criando o cabeçalho na planilha
        for i in self.cabecalho:
            self.ws[i] = self.cabecalho[i]

        # Mesclando células para aplicar os tópicos
        self.ws.merge_cells("A10:A11")
        self.ws.merge_cells("B10:B11")
        self.ws.merge_cells("C10:C11")
        self.ws.merge_cells("D10:D11")
        self.ws.merge_cells("E10:E11")
        self.ws.merge_cells("F10:F11")
        self.ws.merge_cells("G10:G11")
        self.ws.merge_cells("H10:H11")

        # Adicionando os tópicos às células mescladas
        tabela_planilha = tabela.drop(columns=["Código de barras"], inplace=False)

        for i in self.colunas:
            self.ws[i] = self.colunas[i]

        # Adicionando dados do DataFrame na planilha
        for i in range(len(tabela_planilha)):
            for j in range(len(tabela_planilha.columns)):
                self.ws.cell(row=i+12, column=j+1).value = tabela_planilha.iloc[i, j]

        # Definindo variáveis de estilos de fonte da planilha
        fonte_titulo = Font(name="Verdana", size=10)
        fonte_dados = Font(name="Verdana", size=8)

        # Definindo variáveis de alinhamento da planilha
        alinhamento_titulo = Alignment(horizontal="center", vertical="center")
        alinhamento_dados = Alignment(horizontal="left", vertical="center")
        alinhamento_rotulo_cabecalho = Alignment(horizontal="right", vertical="center")
        alinhamento_dados_cabecalho = Alignment(horizontal="center", vertical="center")

        # Definindo variáveis de borda da planilha
        intervalo = self.ws["A1:A5"]
        titulo_celulas = self.ws["A10:H11"]

        # Aplicando as bordas da planilha ao cabeçalho
        intervalo = self.ws["A1:H9"]
        a = [self.ws["A1"], self.ws["A2"], self.ws["A3"], self.ws["A4"], self.ws["A5"], self.ws["A6"], self.ws["A7"], self.ws["A8"], self.ws["A9"]]
        h = [self.ws["H1"], self.ws["H2"], self.ws["H3"], self.ws["H4"], self.ws["H5"], self.ws["H6"], self.ws["H7"], self.ws["H8"], self.ws["H9"]]
        primeira_linha = [self.ws["A1"], self.ws["B1"], self.ws["C1"], self.ws["D1"], self.ws["E1"], self.ws["F1"], self.ws["G1"], self.ws["H1"]]
        ultima_linha = [self.ws["A9"], self.ws["B9"], self.ws["C9"], self.ws["D9"], self.ws["E9"], self.ws["F9"], self.ws["G9"], self.ws["H9"]]

        for i in intervalo:
            for j in i:
                if j == a[0]:
                    j.border = Border(left=Side(style="thin"), top=Side(style="thin"))
                elif j == a[8]:
                    j.border = Border(left=Side(style="thin"), bottom=Side(style="thin"))
                elif j in a:
                    j.border = Border(left=Side(style="thin"))
                elif j == h[0]:
                    j.border = Border(right=Side(style="thin"), top=Side(style="thin"))
                elif j == h[8]:
                    j.border = Border(right=Side(style="thin"), bottom=Side(style="thin"))
                elif j in h:
                    j.border = Border(right=Side(style="thin"))
                elif j in primeira_linha:
                    j.border = Border(top=Side(style="thin"))
                elif j in ultima_linha:
                    j.border = Border(bottom=Side(style="thin"))

        # Aplicando bordas da planilha aos títulos
        celulas_mescladas = [self.ws["A10"], self.ws["B10"], self.ws["C10"], self.ws["D10"], self.ws["E10"], self.ws["F10"], self.ws["G10"], self.ws["H10"]]

        for i in titulo_celulas:
            for j in i:
                if j in celulas_mescladas:
                    j.border = Border(left=Side(style="thin"), top=Side(style="thin"))
                else:
                    j.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        # Aplicando estilos de fonte e alinhamento
        for i in self.cabecalho:
            rotulos = ["A2", "A3", "A4", "E4", "E6"]
            dados = ["B2", "B3", "B4", "F4", "F6"]

            if i in rotulos:
                self.ws[i].alignment = alinhamento_rotulo_cabecalho
                self.ws[i].font = fonte_titulo
            elif i in dados:
                self.ws[i].alignment = alinhamento_dados_cabecalho
                self.ws[i].font = fonte_dados

        for i in self.colunas:
            self.ws[i].alignment = alinhamento_titulo
            self.ws[i].font = fonte_titulo

        for i in range(12, len(tabela_planilha)+12):
            for j in range(1, len(tabela_planilha.columns)+1):
                self.ws.cell(row=i, column=j).alignment = alinhamento_dados
                self.ws.cell(row=i, column=j).font = fonte_dados

    def salvar_planilha(self):
        try:
            self.planilha()
            data = datetime.now(self.fuso_horario).strftime("%d%m%Y")
            hora = datetime.now(self.fuso_horario).strftime("%H%M")
            nome_arquivo = f"{data}_{hora}_relatorio.xlsx"
                
            caminho = os.path.join(os.path.expanduser("~"), "Downloads")
            self.wb.save(os.path.join(caminho, nome_arquivo))

            return True
        except Exception:
            return False
RelatorioExcel().salvar_planilha()